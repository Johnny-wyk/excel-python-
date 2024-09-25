# --coding:utf-8--
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
#数字变成字母
print(get_column_letter(20))

#变成数字
print(column_index_from_string('T'))

#删除工作表
wb=load_workbook('lucky.xlsx')
sheet=wb.get_sheet_names('sheet1')
wb.remove('sheet')

wb.save('lucky.xlsx')