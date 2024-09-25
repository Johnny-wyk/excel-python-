# --coding:utf-8--

from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment

wb=load_workbook('lucky.xlsx')

ws=wb.active

font=Font(size=24,bold=True,color=colors.BLUE)

ws['A1'].font=font

#设置对齐方式
ws['B1'].alignment=Alignment(horizontal='center',vertical='center')
wb.save('lucky.xlsx')