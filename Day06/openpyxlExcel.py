# --coding:utf-8--

from openpyxl import Workbook

wb=Workbook()

#抓取工作表
ws=wb.active

ws['A1']='A1'

ws.append([1,2,3])

from datetime import datetime
ws['A3']=datetime.now()

wb.save('test.xlsx')