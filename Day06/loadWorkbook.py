# --coding:utf-8--

from openpyxl import load_workbook
from datetime import datetime

wb=load_workbook('test.xlsx')

ws=wb.active

ws['A2']='wang'

# #往下一行写入数据
# ws.append(['se','2334'])
#
# ws['A3']=datetime.now().strftime('%Y-%m-%d')


wb.save('test.xlsx')