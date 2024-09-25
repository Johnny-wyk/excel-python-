# --coding:utf-8--

from openpyxl import load_workbook

wb=load_workbook('test.xlsx')

sheets=wb.sheetnames
print(sheets)

