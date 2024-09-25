# --coding:utf-8--

from openpyxl import load_workbook
wb=load_workbook('lucky.xlsx')

ws=wb.active

c_range=ws['A1:C2']
print(c_range)