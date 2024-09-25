# --coding:utf-8--

from openpyxl import load_workbook

wb=load_workbook('lucky.xlsx')

sheet=wb['Sheet']
print(sheet)

#合并单元格
# sheet.merge_cells('A3:A5')

#拆分单元格
sheet.unmerge_cells('A3:A5')
wb.save('lucky.xlsx')