# --coding:utf-8--

from openpyxl import load_workbook
wb=load_workbook('lucky.xlsx')
#获取工作表
ws=wb.active
ws['A1']='Johnny'

#单元格写入
ws.cell(row=3,column=2,value='123432')

for i in range(1,10):
    for j in range(1,5):
        ws.cell(row=i,column=j,value=i*j)
wb.save('lucky.xlsx')