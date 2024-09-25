# --coding:utf-8--

from openpyxl import load_workbook
wb=load_workbook('lucky.xlsx')

ws=wb.active

for row in ws.iter_rows(min_row=1,max_col=3,max_row=5):
    for col in row:
        print(col.value)


for col in ws.iter_cols(min_row=1,max_col=3,max_row=5):
    for row in col:
        print(row.value)
    print()